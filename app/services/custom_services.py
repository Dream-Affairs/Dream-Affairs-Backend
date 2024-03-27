"""This module contains the custom services for the application."""
import csv
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from sqlalchemy.ext.declarative import DeclarativeMeta


#  not used
def model_to_dict(models: List[DeclarativeMeta]) -> List[Dict[str, Any]]:
    """Converts a list of SQLAlchemy models to a list of dictionaries.

    Args:
        models (List[DeclarativeMeta]): List of SQLAlchemy models

    Returns:
        List[Dict[str, Any]]: A dictionary representations of the models
    """

    def convert_datetime(value: Any) -> Any:
        """Converts datetime objects to ISO format."""
        if isinstance(value, datetime):
            return value.isoformat()
        return value

    result_list: List[Dict[str, Any]] = [
        {
            column.name: convert_datetime(getattr(model, column.name))
            for column in model.__table__.columns
        }
        for model in models
    ]

    return result_list


def generate_xlsx_rows(filename):
    """This function generates rows from an Excel file."""
    # Load the workbook
    wb: Workbook = load_workbook(filename)

    # Select the active worksheet
    ws = wb.active

    # Iterate through rows and yield each row
    for row in ws.iter_rows(values_only=True):
        yield row


def generate_csv_rows(filename):
    """This function generates rows from a CSV file."""
    # Open the CSV file
    with open(filename, "r", newline="", encoding="utf-8") as file:
        reader = csv.DictReader(file)
        # Iterate through rows and yield each row
        for row in reader:
            yield row


# def count_rows(filename):
#     """This function counts the number of rows in a file."""
#     num_rows = 0
#     for _ in generate_rows(filename):
#         num_rows += 1
#     return num_rows


def generate_rows(filename: str, file_type: str):
    """This function generates rows from a file."""
    if file_type == "csv":
        return generate_csv_rows(filename)
    if file_type == "xlsx":
        return generate_xlsx_rows(filename)
    raise ValueError(f"Unsupported file type: {file_type}")
